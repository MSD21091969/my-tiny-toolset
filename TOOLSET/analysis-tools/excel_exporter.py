"""
Excel Exporter - Export code analysis to Excel with formatting
Requires: openpyxl
"""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export not available.")
    print("Install with: pip install openpyxl")

from datetime import datetime
from pathlib import Path
from typing import List
from code_analyzer import CodeAnalyzer, ModelInfo, FunctionInfo, RequestResponseMapping


class ExcelExporter:
    """Export code analysis to formatted Excel workbook"""

    def __init__(self, analyzer: CodeAnalyzer):
        self.analyzer = analyzer
        self.wb = None

    def export(self, output_file: str = None) -> str:
        """Export to Excel file"""
        if not EXCEL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. Install with: pip install openpyxl"
            )

        if output_file is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_file = f".tool-outputs/excel/code_analysis_{timestamp}.xlsx"
        
        # Ensure output directory exists
        Path(output_file).parent.mkdir(parents=True, exist_ok=True)

        self.wb = Workbook()

        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])

        # Create sheets
        self._create_summary_sheet()
        self._create_models_sheet()
        self._create_model_fields_sheet()
        self._create_functions_sheet()
        self._create_mappings_sheet()

        # Save workbook
        self.wb.save(output_file)
        print(f"Ô£ô Exported to Excel: {output_file}")
        return output_file

    def _create_summary_sheet(self):
        """Create summary overview sheet"""
        ws = self.wb.create_sheet("Summary", 0)

        # Title
        ws["A1"] = "Code Analysis Summary"
        ws["A1"].font = Font(size=16, bold=True)

        # Basic info
        row = 3
        ws[f"A{row}"] = "Analysis Date:"
        ws[f"B{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        row += 1

        ws[f"A{row}"] = "Root Path:"
        ws[f"B{row}"] = str(self.analyzer.root_path)
        row += 2

        # Statistics
        ws[f"A{row}"] = "Statistics"
        ws[f"A{row}"].font = Font(size=14, bold=True)
        row += 1

        stats = [
            ("Total Models", len(self.analyzer.models)),
            ("Pydantic Models", sum(1 for m in self.analyzer.models if m.is_pydantic)),
            ("Dataclasses", sum(1 for m in self.analyzer.models if m.is_dataclass)),
            ("", ""),
            ("Total Functions", len(self.analyzer.functions)),
            ("Async Functions", sum(1 for f in self.analyzer.functions if f.is_async)),
            ("", ""),
            ("Request/Response Mappings", len(self.analyzer.mappings)),
        ]

        for label, value in stats:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            if label:
                ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Format columns
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 50

    def _create_models_sheet(self):
        """Create models overview sheet"""
        ws = self.wb.create_sheet("Models")

        # Headers
        headers = [
            "Model Name",
            "File Path",
            "Line",
            "Base Classes",
            "Fields",
            "Pydantic",
            "Dataclass",
            "Decorators",
            "Description",
        ]
        self._write_header_row(ws, headers)

        # Data
        for row_idx, model in enumerate(self.analyzer.models, start=2):
            ws[f"A{row_idx}"] = model.name
            ws[f"B{row_idx}"] = model.file_path
            ws[f"C{row_idx}"] = model.line_number
            ws[f"D{row_idx}"] = ", ".join(model.base_classes)
            ws[f"E{row_idx}"] = len(model.fields)
            ws[f"F{row_idx}"] = "Yes" if model.is_pydantic else "No"
            ws[f"G{row_idx}"] = "Yes" if model.is_dataclass else "No"
            ws[f"H{row_idx}"] = ", ".join(model.decorators)
            ws[f"I{row_idx}"] = (model.docstring or "")[:100]

            # Highlight Pydantic models
            if model.is_pydantic:
                ws[f"F{row_idx}"].fill = PatternFill(
                    start_color="90EE90", end_color="90EE90", fill_type="solid"
                )

        # Format columns
        self._auto_adjust_columns(ws)

    def _create_model_fields_sheet(self):
        """Create detailed model fields sheet"""
        ws = self.wb.create_sheet("Model Fields")

        # Headers
        headers = [
            "Model Name",
            "File Path",
            "Field Name",
            "Field Type",
            "Default Value",
            "Optional",
        ]
        self._write_header_row(ws, headers)

        # Data
        row_idx = 2
        for model in self.analyzer.models:
            for field in model.fields:
                ws[f"A{row_idx}"] = model.name
                ws[f"B{row_idx}"] = model.file_path
                ws[f"C{row_idx}"] = field["name"]
                ws[f"D{row_idx}"] = field["type"]
                ws[f"E{row_idx}"] = field.get("default", "")
                ws[f"F{row_idx}"] = "Yes" if "Optional" in field["type"] else "No"
                row_idx += 1

        # Format columns
        self._auto_adjust_columns(ws)

    def _create_functions_sheet(self):
        """Create functions overview sheet"""
        ws = self.wb.create_sheet("Functions")

        # Headers
        headers = [
            "Function Name",
            "File Path",
            "Line",
            "Class",
            "Parameters",
            "Return Type",
            "Async",
            "Decorators",
            "Description",
        ]
        self._write_header_row(ws, headers)

        # Data
        for row_idx, func in enumerate(self.analyzer.functions, start=2):
            ws[f"A{row_idx}"] = func.name
            ws[f"B{row_idx}"] = func.file_path
            ws[f"C{row_idx}"] = func.line_number
            ws[f"D{row_idx}"] = func.class_name or ""

            # Format parameters
            params = ", ".join(
                [f"{p['name']}: {p.get('type', 'Any')}" for p in func.parameters]
            )
            ws[f"E{row_idx}"] = params

            ws[f"F{row_idx}"] = func.return_type or ""
            ws[f"G{row_idx}"] = "Yes" if func.is_async else "No"
            ws[f"H{row_idx}"] = ", ".join(func.decorators)
            ws[f"I{row_idx}"] = (func.docstring or "")[:100]

            # Highlight async functions
            if func.is_async:
                ws[f"G{row_idx}"].fill = PatternFill(
                    start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"
                )

        # Format columns
        self._auto_adjust_columns(ws)

    def _create_mappings_sheet(self):
        """Create request/response mappings sheet"""
        ws = self.wb.create_sheet("API Mappings")

        # Headers
        headers = [
            "Function Name",
            "File Path",
            "HTTP Method",
            "Endpoint",
            "Request Models",
            "Response Models",
        ]
        self._write_header_row(ws, headers)

        # Data
        for row_idx, mapping in enumerate(self.analyzer.mappings, start=2):
            ws[f"A{row_idx}"] = mapping.function_name
            ws[f"B{row_idx}"] = mapping.file_path
            ws[f"C{row_idx}"] = mapping.http_method or ""
            ws[f"D{row_idx}"] = mapping.endpoint or ""
            ws[f"E{row_idx}"] = ", ".join(mapping.request_models)
            ws[f"F{row_idx}"] = ", ".join(mapping.response_models)

            # Color code by HTTP method
            method_colors = {
                "GET": "B0E0E6",
                "POST": "90EE90",
                "PUT": "FFD700",
                "PATCH": "FFA500",
                "DELETE": "FFB6C1",
            }
            if mapping.http_method and mapping.http_method in method_colors:
                color = method_colors[mapping.http_method]
                ws[f"C{row_idx}"].fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

        # Format columns
        self._auto_adjust_columns(ws)

    def _write_header_row(self, ws, headers: List[str]):
        """Write and format header row"""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_adjust_columns(self, ws, max_width: int = 50):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, max_width)
            ws.column_dimensions[column_letter].width = adjusted_width


def export_to_excel(analyzer: CodeAnalyzer, output_file: str = None) -> str:
    """Helper function to export analyzer results to Excel"""
    exporter = ExcelExporter(analyzer)
    return exporter.export(output_file)


if __name__ == "__main__":
    import argparse
    from code_analyzer import CodeAnalyzer
    
    parser = argparse.ArgumentParser(description="Export code analysis to Excel")
    parser.add_argument("path", nargs="?", default=".", help="Root directory to analyze")
    parser.add_argument("--output", "-o", help="Output Excel file path")
    args = parser.parse_args()
    
    if not EXCEL_AVAILABLE:
        print("Excel export not available. Install openpyxl: pip install openpyxl")
        exit(1)
    
    analyzer = CodeAnalyzer(args.path)
    analyzer.analyze_directory()
    export_to_excel(analyzer, args.output)
